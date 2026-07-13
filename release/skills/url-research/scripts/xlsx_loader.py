"""
XLSX Loader · Excel 记录加载器
从 .xlsx 文件中提取「主题 + URL」记录列表，支持批量模式。

用途: url-research 技能 · 步骤0（新增前置步骤）
输入: str Excel文件路径
输出: list[dict{topic, url}] 结构化记录列表

支持格式:
  - 列名自动匹配：主题/topic/标题、URL/链接/地址/网址
  - 非空校验：主题和URL均为非空才纳入
  - URL 自动补全：缺失协议时补充 https://
"""

import sys
import os
import json
from typing import Optional

# 可选依赖
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def normalize_url(url: str) -> str:
    """URL 标准化：去除空白、补充协议前缀"""
    url = url.strip()
    if not url:
        return ""
    if not url.startswith("http://") and not url.startswith("https://"):
        url = "https://" + url
    return url


def is_valid_url(url: str) -> bool:
    """基础 URL 格式校验"""
    url = url.strip()
    if not url:
        return False
    if not url.startswith("http://") and not url.startswith("https://"):
        url = "https://" + url
    try:
        from urllib.parse import urlparse
        result = urlparse(url)
        return all([result.scheme in ("http", "https"), result.netloc])
    except Exception:
        return False


def detect_columns(headers: list[str]) -> tuple[Optional[str], Optional[str]]:
    """
    自动识别主题列和 URL 列。
    返回 (topic_key, url_key)，未识别到则对应位置为 None。
    """
    topic_key = None
    url_key = None

    topic_patterns = ["主题", "topic", "标题", "title", "名称", "name", "话题"]
    url_patterns = ["url", "链接", "地址", "网址", "link", "来源"]

    for h in headers:
        h_lower = h.lower().strip()
        if topic_key is None and any(p in h_lower for p in topic_patterns):
            topic_key = h
        if url_key is None and any(p in h_lower for p in url_patterns):
            url_key = h

    # 回退：第一列当主题，第二列当 URL
    if topic_key is None and len(headers) >= 1:
        topic_key = headers[0]
    if url_key is None and len(headers) >= 2:
        url_key = headers[1]

    return topic_key, url_key


def load_xlsx(file_path: str, sheet_name: Optional[str] = None) -> list[dict]:
    """
    加载 Excel 文件并提取「主题 + URL」记录。

    Args:
        file_path: .xlsx 文件路径
        sheet_name: 指定 sheet 名称，默认第一个 sheet

    Returns:
        list[dict{topic, url}]

    Raises:
        FileNotFoundError: 文件不存在
        ImportError: openpyxl 未安装
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "缺少 openpyxl 依赖，请执行: pip install openpyxl\n"
            "或: pip install -r requirements.txt"
        )

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    # 首行为列名
    headers = [str(h).strip() if h else "" for h in rows[0]]
    topic_key, url_key = detect_columns(headers)

    if not topic_key or not url_key:
        return []

    topic_idx = headers.index(topic_key)
    url_idx = headers.index(url_key)

    records = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        topic = str(row[topic_idx]).strip() if topic_idx < len(row) and row[topic_idx] else ""
        url = str(row[url_idx]).strip() if url_idx < len(row) and row[url_idx] else ""

        if not topic or not url:
            continue

        url = normalize_url(url)
        if not is_valid_url(url):
            continue

        records.append({"topic": topic, "url": url})

    return records


def load(file_path: str, sheet_name: Optional[str] = None) -> dict:
    """
    主入口：加载 Excel 并返回完整的结构化结果。

    Returns:
        {
            "file": str,
            "sheet": str,
            "total_records": int,
            "valid_records": int,
            "skipped_records": int,
            "records": list[dict{topic, url}],
            "urls": list[str],            # 所有URL（供后续步骤直接使用）
            "topic_groups": dict,         # 按 topic 分组的 URL 映射
        }
    """
    records = load_xlsx(file_path, sheet_name)
    urls = list(set(r["url"] for r in records))

    topic_groups = {}
    for r in records:
        t = r["topic"]
        if t not in topic_groups:
            topic_groups[t] = []
        topic_groups[t].append(r["url"])

    return {
        "file": os.path.basename(file_path),
        "sheet": sheet_name or "Sheet1",
        "total_records": len(records),
        "valid_records": len(urls),
        "skipped_records": 0,
        "records": records,
        "urls": urls,
        "topic_groups": topic_groups,
        "topics": list(topic_groups.keys()),
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: python xlsx_loader.py <file_path> [sheet_name]"}, ensure_ascii=False))
        sys.exit(1)

    file_path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        result = load(file_path, sheet)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(1)
